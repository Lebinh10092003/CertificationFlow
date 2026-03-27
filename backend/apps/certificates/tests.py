from io import BytesIO
from unittest.mock import patch

import fitz
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.certificates.models import CertificateExtraction, CertificateMatch, CertificatePage, SourcePdfBatch
from apps.certificates.services.delivery import ensure_public_identity
from apps.certificates.services.exporting import build_batch_export_workbook
from apps.certificates.services.matching import match_certificate_page
from apps.certificates.services.parsing import parse_certificate_text
from apps.certificates.services.pipeline import analyze_source_batch, split_and_process_source_batch
from apps.competitions.models import Competition, IntegrationConfig
from apps.participants.models import CompetitionEnrollment, CompetitionResult, Participant


def build_test_pdf() -> bytes:
    document = fitz.open()
    page = document.new_page()
    page.insert_text(
        (72, 72),
        "Math Olympiad 2026\nCertificate\nNguyen Van A\nGold Medal\nMATH-G-001",
        fontsize=14,
    )
    return document.tobytes()


class PdfPipelineTests(TestCase):
    def test_split_extract_and_match_batch(self):
        competition = Competition.objects.create(name="Math Olympiad 2026", academic_year="2025-2026", subject="Mathematics")
        IntegrationConfig.objects.create(competition=competition)
        participant = Participant.objects.create(
            external_student_id="ST001",
            full_name="Nguyen Van A",
            normalized_name="nguyen van a",
            email="a@example.com",
            school_name="Le Hong Phong",
            normalized_school_name="le hong phong",
            grade="12",
        )
        enrollment = CompetitionEnrollment.objects.create(competition=competition, participant=participant, source_row_number=2, subject="Mathematics")
        CompetitionResult.objects.create(
            competition_enrollment=enrollment,
            award="Gold Medal",
            certificate_code="MATH-G-001",
        )

        batch = SourcePdfBatch.objects.create(
            competition=competition,
            original_filename="certificates.pdf",
            confirmed_competition_name=competition.name,
            competition_confirmation_status=SourcePdfBatch.ConfirmationStatus.CONFIRMED,
            status=SourcePdfBatch.Status.READY,
            processing_mode=SourcePdfBatch.ProcessingMode.SPLIT_EXTRACT_MATCH,
        )
        batch.uploaded_file.save("certificates.pdf", ContentFile(build_test_pdf()), save=True)

        analyze_source_batch(batch)
        split_and_process_source_batch(batch)

        self.assertEqual(CertificatePage.objects.count(), 1)
        page = CertificatePage.objects.select_related("extraction", "match").get()
        self.assertEqual(page.extraction.certificate_code, "MATH-G-001")
        self.assertEqual(page.match.confidence_label, CertificateMatch.ConfidenceLabel.HIGH)
        self.assertEqual(page.match.competition_enrollment, enrollment)

    def test_parse_certificate_text_prefers_awarded_to_real_code_and_grade(self):
        parsed = parse_certificate_text(
            "\n".join(
                [
                    "CERTIFICATE OF EXCELLENCE",
                    "SCO International Coding Olympiad",
                    "This recognises your participation in SCO International School Connect Olympiad 2026-27",
                    "Supported BY",
                    "Sufia Kausher",
                    "Content Head and Co - Founder",
                    "School Name - Truong Tieu hoc Ha Long",
                    "Awarded to Nguyen Hoang Bach",
                    "Grade - 3 - Achieved - Bronze",
                    "SCO Olympiad, 2026",
                    "SCO-IN- OLY202627CYL14101",
                    "Qualified International Round - Yes",
                ]
            )
        )

        self.assertEqual(parsed["student_name"], "Nguyen Hoang Bach")
        self.assertEqual(parsed["school_name"], "Truong Tieu hoc Ha Long")
        self.assertEqual(parsed["grade"], "3")
        self.assertEqual(parsed["award"], "Bronze")
        self.assertEqual(parsed["subject"], "ICO")
        self.assertEqual(parsed["certificate_code"], "SCO-IN-OLY202627CYL14101")
        self.assertEqual(parsed["qualified_round"], "Yes")

    def test_parse_certificate_text_handles_sco_edited_pdf_format(self):
        parsed = parse_certificate_text(
            "\n".join(
                [
                    "CERTIFICATE OF EXCELLENCE",
                    "SCO International Coding Olympiad",
                    "SCO-IN- OLY202627CYL14101",
                    "This recognises your participation in SCO International School Connect Olympiad 2026-27",
                    "Supported BY",
                    "Awarded to Nguyễn Hoàng Bách",
                    "Grade 3 - Achieved - Bronze Medal",
                    "School Name - Trường Tiểu học Hạ Long",
                    "Qualified International Round - Yes",
                ]
            )
        )

        self.assertEqual(parsed["competition_name"], "SCO International Coding Olympiad")
        self.assertEqual(parsed["student_name"], "Nguyễn Hoàng Bách")
        self.assertEqual(parsed["normalized_student_name"], "nguyen hoang bach")
        self.assertEqual(parsed["school_name"], "Trường Tiểu học Hạ Long")
        self.assertEqual(parsed["grade"], "3")
        self.assertEqual(parsed["award"], "Bronze Medal")
        self.assertEqual(parsed["subject"], "ICO")
        self.assertEqual(parsed["certificate_code"], "SCO-IN-OLY202627CYL14101")
        self.assertEqual(parsed["qualified_round"], "Yes")

    def test_match_certificate_page_promotes_unique_name_grade_competition_to_high(self):
        competition = Competition.objects.create(name="SCO International Coding Olympiad", academic_year="2026-2027", subject="ICO")
        participant = Participant.objects.create(
            external_student_id="ST-001",
            full_name="Nguyễn Hoàng Bách",
            normalized_name="nguyen hoang bach",
            email="bach@example.com",
            school_name="Trường Tiểu học Hạ Long",
            normalized_school_name="truong tieu hoc ha long",
            grade="3",
        )
        enrollment = CompetitionEnrollment.objects.create(
            competition=competition,
            participant=participant,
            subject="ICO",
            source_row_number=2,
        )
        result = CompetitionResult.objects.create(
            competition_enrollment=enrollment,
            award="Bronze",
            qualified_round="Yes",
        )
        batch = SourcePdfBatch.objects.create(
            competition=competition,
            original_filename="certificates.pdf",
            confirmed_competition_name=competition.name,
            competition_confirmation_status=SourcePdfBatch.ConfirmationStatus.CONFIRMED,
            status=SourcePdfBatch.Status.COMPLETED,
        )
        batch.uploaded_file.save("certificates.pdf", ContentFile(build_test_pdf()), save=True)
        page = CertificatePage.objects.create(
            source_batch=batch,
            page_number=1,
            processing_status=CertificatePage.ProcessingStatus.EXTRACTED,
        )
        CertificateExtraction.objects.create(
            certificate_page=page,
            student_name="Nguyễn Hoàng Bách",
            normalized_student_name="nguyen hoang bach",
            school_name="Trường Tiểu học Hạ Long",
            normalized_school_name="truong tieu hoc ha long",
            grade="3",
            award="Bronze Medal",
            subject="ICO",
            competition_name=competition.name,
            certificate_code="",
            qualified_round="Yes",
        )

        match = match_certificate_page(page)

        self.assertEqual(match.competition_enrollment, enrollment)
        self.assertEqual(match.competition_result, result)
        self.assertEqual(match.confidence_label, CertificateMatch.ConfidenceLabel.HIGH)
        self.assertFalse(match.requires_review)

    @patch("apps.certificates.views.analyze_source_batch_task.delay")
    def test_upload_batch_attaches_selected_competition(self, mock_delay):
        competition = Competition.objects.create(name="Attached Competition", academic_year="2026-2027")
        client = APIClient()
        user = get_user_model().objects.create_user(username="uploadtester", password="testpass123")
        client.force_authenticate(user)

        response = client.post(
            "/api/certificate-batches/",
            {
                "file": SimpleUploadedFile("attached.pdf", build_test_pdf(), content_type="application/pdf"),
                "processing_mode": SourcePdfBatch.ProcessingMode.SPLIT_EXTRACT_MATCH,
                "competition_id": competition.id,
            },
        )

        self.assertEqual(response.status_code, 201)
        batch = SourcePdfBatch.objects.get(pk=response.json()["id"])
        self.assertEqual(batch.competition_id, competition.id)
        filtered_response = client.get(f"/api/certificate-batches/?competition={competition.id}")
        self.assertEqual(filtered_response.status_code, 200)
        self.assertEqual(len(filtered_response.json()), 1)
        mock_delay.assert_called_once_with(batch.id)


class CertificateDeliveryTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(username="deliverytester", password="testpass123")
        self.client.force_authenticate(self.user)
        self.competition = Competition.objects.create(
            name="SCO International Coding Olympiad",
            academic_year="2026-2027",
            subject="ICO",
        )
        self.integration = IntegrationConfig.objects.create(
            competition=self.competition,
            sheets_spreadsheet_id="sheet123",
            sheets_worksheet_name="ICO_OUTPUT",
            drive_folder_id="folder123",
        )
        self.participant = Participant.objects.create(
            external_student_id="ST001",
            full_name="Nguyen Hoang Bach",
            normalized_name="nguyen hoang bach",
            email="bach@example.com",
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade="3",
        )
        self.enrollment = CompetitionEnrollment.objects.create(
            competition=self.competition,
            participant=self.participant,
            source_row_number=2,
            subject="ICO",
            sheet_row_key="2",
            source_data_json={
                "No.": "1",
                "Candidate's Name": "Nguyen Hoang Bach",
                "School/ Institution Name": "Ha Long Primary School",
                "Grade": "3",
                "Competition": "ICO",
                "Parent Email": "bach@example.com",
                "Award Status": "Bronze",
                "Qualification Status": "Yes",
            },
        )
        self.result = CompetitionResult.objects.create(
            competition_enrollment=self.enrollment,
            award="Bronze",
            certificate_code="SCO-IN-OLY202627CYL14101",
            qualified_round="Yes",
        )
        self.batch = SourcePdfBatch.objects.create(
            competition=self.competition,
            original_filename="sco_batch.pdf",
            confirmed_competition_name=self.competition.name,
            competition_confirmation_status=SourcePdfBatch.ConfirmationStatus.CONFIRMED,
            status=SourcePdfBatch.Status.COMPLETED,
        )
        self.batch.uploaded_file.save("sco_batch.pdf", ContentFile(build_test_pdf()), save=True)
        self.page = CertificatePage.objects.create(
            source_batch=self.batch,
            page_number=1,
            processing_status=CertificatePage.ProcessingStatus.MATCHED,
            output_filename="nguyen-hoang-bach-bronze-ico.pdf",
        )
        self.page.split_pdf_file.save("page_1.pdf", ContentFile(build_test_pdf()), save=True)
        CertificateExtraction.objects.create(
            certificate_page=self.page,
            student_name="Nguyen Hoang Bach",
            normalized_student_name="nguyen hoang bach",
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade="3",
            award="Bronze",
            subject="ICO",
            competition_name=self.competition.name,
            certificate_code="SCO-IN-OLY202627CYL14101",
            qualified_round="Yes",
        )
        CertificateMatch.objects.create(
            certificate_page=self.page,
            competition_enrollment=self.enrollment,
            competition_result=self.result,
            confidence_score=100,
            confidence_label=CertificateMatch.ConfidenceLabel.HIGH,
            matched_by=CertificateMatch.MatchedBy.CERTIFICATE_CODE,
            requires_review=False,
            is_approved=True,
            rationale="Exact code",
        )

    def _create_second_page(self, *, sheet_name: str, row_number: int, student_name: str):
        participant = Participant.objects.create(
            external_student_id=f"ST{row_number:03d}",
            full_name=student_name,
            normalized_name=student_name.lower(),
            email=f"{student_name.split()[0].lower()}@example.com",
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade="4",
        )
        enrollment = CompetitionEnrollment.objects.create(
            competition=self.competition,
            participant=participant,
            source_row_number=row_number,
            source_sheet_name=sheet_name,
            subject="ICO",
            sheet_row_key=str(row_number),
            source_data_json={"Candidate's Name": student_name},
        )
        result = CompetitionResult.objects.create(
            competition_enrollment=enrollment,
            award="Silver",
            certificate_code=f"SCO-IN-TEST-{row_number}",
            qualified_round="Yes",
        )
        page = CertificatePage.objects.create(
            source_batch=self.batch,
            page_number=row_number,
            processing_status=CertificatePage.ProcessingStatus.MATCHED,
            output_filename=f"{student_name.lower().replace(' ', '-')}.pdf",
        )
        page.split_pdf_file.save(f"page_{row_number}.pdf", ContentFile(build_test_pdf()), save=True)
        CertificateExtraction.objects.create(
            certificate_page=page,
            student_name=student_name,
            normalized_student_name=student_name.lower(),
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade="4",
            award="Silver",
            subject="ICO",
            competition_name=self.competition.name,
            certificate_code=f"SCO-IN-TEST-{row_number}",
            qualified_round="Yes",
        )
        CertificateMatch.objects.create(
            certificate_page=page,
            competition_enrollment=enrollment,
            competition_result=result,
            confidence_score=100,
            confidence_label=CertificateMatch.ConfidenceLabel.HIGH,
            matched_by=CertificateMatch.MatchedBy.CERTIFICATE_CODE,
            requires_review=False,
            is_approved=True,
            rationale="Exact code",
        )
        return page

    def _create_additional_batch_with_page(self, *, filename: str, student_name: str, award: str, code: str, grade: str):
        participant = Participant.objects.create(
            external_student_id=f"ST-{code}",
            full_name=student_name,
            normalized_name=student_name.lower(),
            email=f"{student_name.split()[0].lower()}@example.com",
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade=grade,
        )
        enrollment = CompetitionEnrollment.objects.create(
            competition=self.competition,
            participant=participant,
            source_row_number=2,
            source_sheet_name="ICO_OUTPUT",
            subject="ICO",
            sheet_row_key="2",
            source_data_json={
                "Candidate's Name": student_name,
                "Grade": grade,
                "Competition": "ICO",
                "Award Status": award,
            },
        )
        result = CompetitionResult.objects.create(
            competition_enrollment=enrollment,
            award=award,
            certificate_code=code,
            qualified_round="Yes",
        )
        batch = SourcePdfBatch.objects.create(
            competition=self.competition,
            original_filename=filename,
            confirmed_competition_name=self.competition.name,
            competition_confirmation_status=SourcePdfBatch.ConfirmationStatus.CONFIRMED,
            status=SourcePdfBatch.Status.COMPLETED,
        )
        batch.uploaded_file.save(filename, ContentFile(build_test_pdf()), save=True)
        page = CertificatePage.objects.create(
            source_batch=batch,
            page_number=1,
            processing_status=CertificatePage.ProcessingStatus.MATCHED,
            output_filename=f"{student_name.lower().replace(' ', '-')}.pdf",
        )
        page.split_pdf_file.save("page_1.pdf", ContentFile(build_test_pdf()), save=True)
        CertificateExtraction.objects.create(
            certificate_page=page,
            student_name=student_name,
            normalized_student_name=student_name.lower(),
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade=grade,
            award=award,
            subject="ICO",
            competition_name=self.competition.name,
            certificate_code=code,
            qualified_round="Yes",
        )
        CertificateMatch.objects.create(
            certificate_page=page,
            competition_enrollment=enrollment,
            competition_result=result,
            confidence_score=100,
            confidence_label=CertificateMatch.ConfidenceLabel.HIGH,
            matched_by=CertificateMatch.MatchedBy.CERTIFICATE_CODE,
            requires_review=False,
            is_approved=True,
            rationale="Exact code",
        )
        return batch, page

    def test_ensure_public_identity_generates_slug_with_grade(self):
        ensure_public_identity(self.page)
        self.page.refresh_from_db()

        self.assertEqual(self.page.public_slug, "nguyen-hoang-bach-bronze-ico-3")
        self.assertTrue(self.page.public_url.endswith("/c/nguyen-hoang-bach-bronze-ico-3"))

    def test_build_batch_export_workbook_adds_public_link_by_default(self):
        ensure_public_identity(self.page)

        filename, workbook_bytes = build_batch_export_workbook(self.batch)

        self.assertEqual(filename, f"batch_{self.batch.id}_delivery.xlsx")
        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook["ICO_OUTPUT"]
        headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        row_values = {header: sheet.cell(row=2, column=index).value for index, header in enumerate(headers, start=1)}

        self.assertIn("Public Link", headers)
        self.assertEqual(row_values["Candidate's Name"], "Nguyen Hoang Bach")
        self.assertEqual(row_values["Grade"], "3")
        self.assertTrue(row_values["Public Link"].endswith("/c/nguyen-hoang-bach-bronze-ico-3"))
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertTrue(bool(sheet.auto_filter.ref))
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "001F4E78")

    def test_build_batch_export_workbook_excludes_unapproved_pages(self):
        self.page.match.is_approved = False
        self.page.match.save(update_fields=["is_approved", "updated_at"])

        _filename, workbook_bytes = build_batch_export_workbook(self.batch)

        workbook = load_workbook(BytesIO(workbook_bytes))
        sheet = workbook[workbook.sheetnames[0]]
        self.assertEqual(sheet.max_row, 1)

    def test_export_columns_endpoint_lists_source_and_system_columns(self):
        response = self.client.get(f"/api/certificate-batches/{self.batch.id}/export-columns/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(any(column["key"] == "Candidate's Name" for column in payload["source_columns"]))
        self.assertTrue(any(column["key"] == "public_link" for column in payload["system_columns"]))
        self.assertFalse(any(column["key"] == "drive_link" for column in payload["system_columns"]))

    def test_configurable_export_endpoint_respects_column_order_and_labels(self):
        ensure_public_identity(self.page)

        response = self.client.post(
            f"/api/certificate-batches/{self.batch.id}/export/",
            {
                "columns": [
                    {"key": "matched_student", "label": "Matched Student", "source_type": "system"},
                    {"key": "public_link", "label": "Certificate URL", "source_type": "system"},
                    {"key": "Candidate's Name", "label": "Original Name", "source_type": "source"},
                ],
                "sheet_mode": "single_sheet",
                "format_mode": "presentation",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook[workbook.sheetnames[0]]
        headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
        self.assertEqual(headers, ["Matched Student", "Certificate URL", "Original Name"])
        self.assertEqual(sheet["B2"].hyperlink.target, self.page.public_url or ensure_public_identity(self.page).public_url)
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_export_endpoint_downloads_workbook(self):
        response = self.client.get(f"/api/certificate-batches/{self.batch.id}/export/")

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_endpoint_rejects_batches_without_approved_pages(self):
        self.page.match.is_approved = False
        self.page.match.save(update_fields=["is_approved", "updated_at"])

        response = self.client.get(f"/api/certificate-batches/{self.batch.id}/export/")

        self.assertEqual(response.status_code, 400)
        self.assertIn("No approved certificate pages", response.json()["detail"])

    def test_competition_export_columns_endpoint_supports_multiple_batches(self):
        second_batch, _second_page = self._create_additional_batch_with_page(
            filename="sco_batch_2.pdf",
            student_name="Tran Minh Khoa",
            award="Silver",
            code="SCO-IN-TEST-2",
            grade="4",
        )

        response = self.client.get(
            f"/api/competitions/{self.competition.id}/certificate-export-columns/?batch_ids={self.batch.id},{second_batch.id}"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(any(column["key"] == "Candidate's Name" for column in payload["source_columns"]))
        self.assertTrue(any(column["key"] == "batch_file" for column in payload["system_columns"]))

    def test_competition_export_endpoint_downloads_selected_batches(self):
        second_batch, second_page = self._create_additional_batch_with_page(
            filename="sco_batch_2.pdf",
            student_name="Tran Minh Khoa",
            award="Silver",
            code="SCO-IN-TEST-2",
            grade="4",
        )
        ensure_public_identity(self.page)
        ensure_public_identity(second_page)

        response = self.client.post(
            f"/api/competitions/{self.competition.id}/certificate-export/",
            {
                "batch_ids": [self.batch.id, second_batch.id],
                "columns": [
                    {"key": "Candidate's Name", "label": "Candidate's Name", "source_type": "source"},
                    {"key": "batch_file", "label": "Batch File", "source_type": "system"},
                    {"key": "public_link", "label": "Public Link", "source_type": "system"},
                ],
                "sheet_mode": "single_sheet",
                "format_mode": "business",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook[workbook.sheetnames[0]]
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet["B2"].value, "sco_batch.pdf")
        self.assertEqual(sheet["B3"].value, "sco_batch_2.pdf")

    def test_competition_export_endpoint_accepts_payload_json_form_post(self):
        ensure_public_identity(self.page)

        response = self.client.post(
            f"/api/competitions/{self.competition.id}/certificate-export/",
            {
                "payload_json": '{"batch_ids":[%d],"columns":[{"key":"public_link","label":"Public Link","source_type":"system"}],"sheet_mode":"single_sheet","format_mode":"business"}'
                % self.batch.id
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_approve_endpoint_generates_public_link(self):
        self.page.match.is_approved = False
        self.page.match.requires_review = True
        self.page.match.save(update_fields=["is_approved", "requires_review", "updated_at"])

        response = self.client.post(f"/api/certificate-matches/{self.page.match.id}/approve/")

        self.assertEqual(response.status_code, 200)
        self.page.refresh_from_db()
        self.page.match.refresh_from_db()
        self.assertTrue(self.page.match.is_approved)
        self.assertFalse(self.page.match.requires_review)
        self.assertTrue(self.page.public_slug)
        self.assertTrue(self.page.public_url.endswith(f"/c/{self.page.public_slug}"))

    def test_bulk_review_endpoint_updates_multiple_matches(self):
        self.page.match.is_approved = False
        self.page.match.requires_review = True
        self.page.match.save(update_fields=["is_approved", "requires_review", "updated_at"])
        second_page = self._create_second_page(
            sheet_name="ICO_OUTPUT",
            row_number=2,
            student_name="Tran Minh Khoa",
        )
        second_page.match.is_approved = False
        second_page.match.requires_review = True
        second_page.match.save(update_fields=["is_approved", "requires_review", "updated_at"])

        response = self.client.post(
            "/api/certificate-matches/bulk-review/",
            {
                "match_ids": [self.page.match.id, second_page.match.id],
                "approved": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.page.refresh_from_db()
        second_page.refresh_from_db()
        self.assertTrue(self.page.match.is_approved)
        self.assertTrue(second_page.match.is_approved)
        self.assertTrue(self.page.public_url)
        self.assertTrue(second_page.public_url)

    def test_bulk_review_endpoint_can_unapprove_matches(self):
        response = self.client.post(
            "/api/certificate-matches/bulk-review/",
            {
                "match_ids": [self.page.match.id],
                "approved": False,
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.page.refresh_from_db()
        self.page.match.refresh_from_db()
        self.assertFalse(self.page.match.is_approved)
        self.assertTrue(self.page.match.requires_review)
        self.assertEqual(self.page.processing_status, CertificatePage.ProcessingStatus.REVIEW_REQUIRED)

    def test_public_certificate_endpoint_returns_public_payload(self):
        ensure_public_identity(self.page)

        response = self.client.get(f"/api/public-certificates/{self.page.public_slug}/")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["student_name"], "Nguyen Hoang Bach")
        self.assertEqual(payload["competition_code"], "ICO")
        self.assertEqual(payload["grade"], "3")
        self.assertTrue(payload["download_pdf_url"].endswith(f"/api/certificate-pages/{self.page.id}/pdf/?download=1"))

    def test_certificate_page_pdf_endpoint_uses_output_filename(self):
        response = self.client.get(f"/api/certificate-pages/{self.page.id}/pdf/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn('filename="nguyen-hoang-bach-bronze-ico.pdf"', response["Content-Disposition"])

    def test_certificate_page_pdf_endpoint_supports_download_attachment(self):
        response = self.client.get(f"/api/certificate-pages/{self.page.id}/pdf/?download=1")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn('filename="nguyen-hoang-bach-bronze-ico.pdf"', response["Content-Disposition"])

    def test_certificate_pages_serializer_returns_pdf_endpoint_url(self):
        response = self.client.get(
            f"/api/certificate-pages/?competition={self.competition.id}&batch={self.batch.id}"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()[0]
        self.assertTrue(payload["split_pdf_url"].endswith(f"/api/certificate-pages/{self.page.id}/pdf/"))
        self.assertTrue(payload["download_pdf_url"].endswith(f"/api/certificate-pages/{self.page.id}/pdf/?download=1"))

    def test_public_certificate_endpoint_hides_unapproved_pages(self):
        ensure_public_identity(self.page)
        self.page.match.is_approved = False
        self.page.match.save(update_fields=["is_approved", "updated_at"])

        response = self.client.get(f"/api/public-certificates/{self.page.public_slug}/")

        self.assertEqual(response.status_code, 404)

    def test_certificate_pages_endpoint_includes_review_status_and_export_ready(self):
        ensure_public_identity(self.page)
        self.page.match.is_approved = False
        self.page.match.save(update_fields=["is_approved", "updated_at"])

        response = self.client.get(
            f"/api/certificate-pages/?competition={self.competition.id}&batch={self.batch.id}"
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()[0]
        self.assertEqual(payload["review_status"], "needs_review")
        self.assertFalse(payload["export_ready"])

    def test_delete_batch_endpoint_keeps_imported_student_data(self):
        response = self.client.delete(f"/api/certificate-batches/{self.batch.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(SourcePdfBatch.objects.filter(pk=self.batch.id).exists())
        self.assertEqual(CertificatePage.objects.count(), 0)
        self.assertEqual(CertificateExtraction.objects.count(), 0)
        self.assertEqual(CertificateMatch.objects.count(), 0)
        self.assertEqual(Participant.objects.count(), 1)
        self.assertEqual(CompetitionEnrollment.objects.count(), 1)
        self.assertEqual(CompetitionResult.objects.count(), 1)
