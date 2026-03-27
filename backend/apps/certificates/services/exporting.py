from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.competitions.models import Competition

from ..models import SourcePdfBatch
from .delivery import batch_delivery_pages, delivery_pages_for_batch_ids, ensure_public_identity, page_competition_code


DEFAULT_EXPORT_HEADERS = [
    "No.",
    "Candidate's Name",
    "School/ Institution Name",
    "Grade",
    "Competition",
    "Parent Email",
    "Teacher Email",
    "Username",
    "Award Status",
    "Qualification Status",
    "Registration Status",
]

SOURCE_TYPE_SOURCE = "source"
SOURCE_TYPE_SYSTEM = "system"
SHEET_MODE_SPLIT = "split_by_competition"
SHEET_MODE_SINGLE = "single_sheet"
FORMAT_MODE_BUSINESS = "business"
FORMAT_MODE_COMPACT = "compact"
FORMAT_MODE_PRESENTATION = "presentation"

SYSTEM_EXPORT_COLUMNS = [
    {"key": "public_link", "label": "Public Link"},
    {"key": "extracted_name", "label": "Extracted Name"},
    {"key": "extracted_school", "label": "Extracted School"},
    {"key": "extracted_grade", "label": "Extracted Grade"},
    {"key": "certificate_code", "label": "Certificate Code"},
    {"key": "matched_student", "label": "Matched Student"},
    {"key": "matched_email", "label": "Matched Email"},
    {"key": "confidence", "label": "Confidence"},
    {"key": "page_number", "label": "Page Number"},
    {"key": "batch_file", "label": "Batch File"},
    {"key": "award", "label": "Award"},
    {"key": "competition", "label": "Competition"},
    {"key": "qualified_round", "label": "Qualified Round"},
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LINK_FONT = Font(name="Calibri", size=11, underline="single", color="0563C1")
BASE_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FBFF")
CENTER_KEYS = {"grade", "confidence", "page_number", "award", "competition", "qualified_round"}
LEFT_KEYS = {"matched_student", "matched_email", "extracted_name", "extracted_school", "certificate_code", "public_link", "batch_file"}
LINK_KEYS = {"public_link"}


def _fallback_source_row(page) -> dict:
    enrollment = page.match.competition_enrollment
    result = page.match.competition_result
    participant = enrollment.participant
    return {
        "No.": str(enrollment.source_row_number or ""),
        "Candidate's Name": participant.full_name,
        "School/ Institution Name": participant.school_name,
        "Grade": participant.grade,
        "Competition": enrollment.subject or page_competition_code(page),
        "Parent Email": participant.email,
        "Teacher Email": "",
        "Username": "",
        "Award Status": result.award if result else page.extraction.award,
        "Qualification Status": result.qualified_round if result else page.extraction.qualified_round,
        "Registration Status": enrollment.notes,
    }


def _source_row(page) -> dict:
    enrollment = page.match.competition_enrollment
    row = {
        key: value
        for key, value in dict(enrollment.source_data_json or {}).items()
        if not str(key).startswith("_")
    }
    if not row:
        return _fallback_source_row(page)
    return row


def _system_row(page, request=None) -> dict:
    ensure_public_identity(page, request=request)
    extraction = page.extraction
    match = page.match
    enrollment = match.competition_enrollment if match else None
    result = match.competition_result if match else None
    participant = enrollment.participant if enrollment else None
    return {
        "public_link": page.public_url,
        "extracted_name": extraction.student_name,
        "extracted_school": extraction.school_name,
        "extracted_grade": extraction.grade,
        "certificate_code": extraction.certificate_code,
        "matched_student": participant.full_name if participant else "",
        "matched_email": participant.email if participant else "",
        "confidence": match.confidence_label if match else "",
        "page_number": page.page_number,
        "batch_file": page.source_batch.original_filename,
        "award": result.award if result and result.award else extraction.award,
        "competition": enrollment.subject if enrollment and enrollment.subject else page_competition_code(page),
        "qualified_round": result.qualified_round if result and result.qualified_round else extraction.qualified_round,
    }


def _default_column_configs(batch: SourcePdfBatch) -> list[dict]:
    source_columns: list[dict] = []
    seen_source_keys: set[str] = set()
    for page in batch_delivery_pages(batch):
        for key in _source_row(page).keys():
            if key in seen_source_keys:
                continue
            seen_source_keys.add(key)
            source_columns.append({"key": key, "label": key, "source_type": SOURCE_TYPE_SOURCE})

    columns = [*source_columns]
    for system_column in SYSTEM_EXPORT_COLUMNS:
        if system_column["key"] == "public_link":
            columns.append({**system_column, "source_type": SOURCE_TYPE_SYSTEM})
    return columns


def _default_column_configs_for_batches(batches: list[SourcePdfBatch]) -> list[dict]:
    source_columns: list[dict] = []
    seen_source_keys: set[str] = set()
    batch_ids = [batch.id for batch in batches if batch.id]
    for page in delivery_pages_for_batch_ids(batch_ids):
        for key in _source_row(page).keys():
            if key in seen_source_keys:
                continue
            seen_source_keys.add(key)
            source_columns.append({"key": key, "label": key, "source_type": SOURCE_TYPE_SOURCE})

    columns = [*source_columns]
    for system_column in SYSTEM_EXPORT_COLUMNS:
        if system_column["key"] == "public_link":
            columns.append({**system_column, "source_type": SOURCE_TYPE_SYSTEM})
    return columns


def get_batch_export_columns(batch: SourcePdfBatch) -> dict:
    source_columns: list[dict] = []
    seen_source_keys: set[str] = set()
    for page in batch_delivery_pages(batch):
        for key in _source_row(page).keys():
            if key in seen_source_keys:
                continue
            seen_source_keys.add(key)
            source_columns.append({"key": key, "label": key, "source_type": SOURCE_TYPE_SOURCE})

    system_columns = [
        {
            **column,
            "source_type": SOURCE_TYPE_SYSTEM,
        }
        for column in SYSTEM_EXPORT_COLUMNS
    ]
    return {
        "source_columns": source_columns,
        "system_columns": system_columns,
        "default_columns": _default_column_configs(batch) if batch.pk else [],
        "sheet_modes": [SHEET_MODE_SPLIT, SHEET_MODE_SINGLE],
        "format_modes": [FORMAT_MODE_BUSINESS, FORMAT_MODE_COMPACT, FORMAT_MODE_PRESENTATION],
    }


def get_batches_export_columns(batches: list[SourcePdfBatch]) -> dict:
    batch_ids = [batch.id for batch in batches if batch.id]
    source_columns: list[dict] = []
    seen_source_keys: set[str] = set()
    for page in delivery_pages_for_batch_ids(batch_ids):
        for key in _source_row(page).keys():
            if key in seen_source_keys:
                continue
            seen_source_keys.add(key)
            source_columns.append({"key": key, "label": key, "source_type": SOURCE_TYPE_SOURCE})

    system_columns = [
        {
            **column,
            "source_type": SOURCE_TYPE_SYSTEM,
        }
        for column in SYSTEM_EXPORT_COLUMNS
    ]
    return {
        "source_columns": source_columns,
        "system_columns": system_columns,
        "default_columns": _default_column_configs_for_batches(batches),
        "sheet_modes": [SHEET_MODE_SPLIT, SHEET_MODE_SINGLE],
        "format_modes": [FORMAT_MODE_BUSINESS, FORMAT_MODE_COMPACT, FORMAT_MODE_PRESENTATION],
    }


def _ordered_columns(columns: list[dict] | None, batch: SourcePdfBatch) -> list[dict]:
    return columns or _default_column_configs(batch)


def _row_value(page, column: dict, request=None):
    if column["source_type"] == SOURCE_TYPE_SOURCE:
        return _source_row(page).get(column["key"], "")
    return _system_row(page, request=request).get(column["key"], "")


def _sheet_title(batch: SourcePdfBatch, competition_code: str, sheet_mode: str) -> str:
    if sheet_mode == SHEET_MODE_SINGLE:
        return f"batch_{batch.id}"[:31]
    return f"{competition_code or 'OUTPUT'}_OUTPUT"[:31]


def _mode_limits(format_mode: str) -> tuple[int, int, float]:
    if format_mode == FORMAT_MODE_COMPACT:
        return 10, 24, 18
    if format_mode == FORMAT_MODE_PRESENTATION:
        return 12, 40, 28
    return 12, 32, 24


def _column_alignment(column: dict) -> Alignment:
    key = column["key"]
    horizontal = "center" if key in CENTER_KEYS else "left"
    if key not in LEFT_KEYS and key not in CENTER_KEYS and len(column.get("label", "")) <= 12:
        horizontal = "center"
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def _apply_sheet_style(sheet, columns: list[dict], format_mode: str):
    min_width, max_width, header_height = _mode_limits(format_mode)
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions

    for column_index, column in enumerate(columns, start=1):
        header_cell = sheet.cell(row=1, column=column_index)
        header_cell.font = HEADER_FONT
        header_cell.fill = HEADER_FILL
        header_cell.border = THIN_BORDER
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = len(str(header_cell.value or ""))

        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = BASE_FONT
            cell.border = THIN_BORDER
            cell.alignment = _column_alignment(column)
            if row_index % 2 == 0:
                cell.fill = ZEBRA_FILL
            cell_value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(cell_value), max_width))
            if column["key"] in LINK_KEYS and cell_value:
                cell.hyperlink = cell_value
                cell.font = LINK_FONT

        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, min_width), max_width)

    sheet.row_dimensions[1].height = header_height


def build_batch_export_workbook(
    batch: SourcePdfBatch,
    *,
    columns: list[dict] | None = None,
    sheet_mode: str = SHEET_MODE_SPLIT,
    format_mode: str = FORMAT_MODE_BUSINESS,
    request=None,
) -> tuple[str, bytes]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    selected_columns = _ordered_columns(columns, batch)

    grouped_pages = defaultdict(list)
    for page in batch_delivery_pages(batch):
        group_key = page_competition_code(page) if sheet_mode == SHEET_MODE_SPLIT else "BATCH"
        grouped_pages[group_key].append(page)

    if not grouped_pages:
        sheet = workbook.create_sheet("OUTPUT")
        for column_index, column in enumerate(selected_columns, start=1):
            sheet.cell(row=1, column=column_index, value=column["label"])
        _apply_sheet_style(sheet, selected_columns, format_mode)
    else:
        for group_key, pages in grouped_pages.items():
            sheet = workbook.create_sheet(_sheet_title(batch, group_key, sheet_mode))
            for column_index, column in enumerate(selected_columns, start=1):
                sheet.cell(row=1, column=column_index, value=column["label"])
            for row_index, page in enumerate(pages, start=2):
                for column_index, column in enumerate(selected_columns, start=1):
                    sheet.cell(row=row_index, column=column_index, value=_row_value(page, column, request=request))
            _apply_sheet_style(sheet, selected_columns, format_mode)

    buffer = BytesIO()
    workbook.save(buffer)
    filename = f"batch_{batch.id}_delivery.xlsx"
    return filename, buffer.getvalue()


def build_batches_export_workbook(
    batches: list[SourcePdfBatch],
    *,
    columns: list[dict] | None = None,
    sheet_mode: str = SHEET_MODE_SPLIT,
    format_mode: str = FORMAT_MODE_BUSINESS,
    request=None,
) -> tuple[str, bytes]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    selected_columns = columns or _default_column_configs_for_batches(batches)
    batch_ids = [batch.id for batch in batches if batch.id]

    grouped_pages = defaultdict(list)
    for page in delivery_pages_for_batch_ids(batch_ids):
        group_key = page_competition_code(page) if sheet_mode == SHEET_MODE_SPLIT else "BATCH"
        grouped_pages[group_key].append(page)

    if not grouped_pages:
        sheet = workbook.create_sheet("OUTPUT")
        for column_index, column in enumerate(selected_columns, start=1):
            sheet.cell(row=1, column=column_index, value=column["label"])
        _apply_sheet_style(sheet, selected_columns, format_mode)
    else:
        for group_key, pages in grouped_pages.items():
            sheet = workbook.create_sheet(_sheet_title(batches[0], group_key, sheet_mode))
            for column_index, column in enumerate(selected_columns, start=1):
                sheet.cell(row=1, column=column_index, value=column["label"])
            for row_index, page in enumerate(pages, start=2):
                for column_index, column in enumerate(selected_columns, start=1):
                    sheet.cell(row=row_index, column=column_index, value=_row_value(page, column, request=request))
            _apply_sheet_style(sheet, selected_columns, format_mode)

    buffer = BytesIO()
    workbook.save(buffer)
    competition = next((batch.competition for batch in batches if batch.competition_id), None)
    if isinstance(competition, Competition):
        filename = f"competition_{competition.id}_certificates.xlsx"
    else:
        filename = "certificates_export.xlsx"
    return filename, buffer.getvalue()
