from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from rest_framework.test import APIClient

from apps.certificates.models import CertificateMatch, CertificatePage, CertificateExtraction, SourcePdfBatch
from apps.competitions.models import Competition
from apps.data_imports.models import DataImportJob
from apps.data_imports.services.tabular import import_rows
from apps.participants.models import CompetitionEnrollment, CompetitionResult, Participant


def build_workbook_bytes(*, include_header: bool = True) -> bytes:
    workbook = Workbook()
    sheet_one = workbook.active
    sheet_one.title = "ICO_OUTPUT"
    sheet_two = workbook.create_sheet("IAIO_OUTPUT")

    if include_header:
        rows = [
            ["Candidate's Name", "School/ Institution Name", "Grade", "Competition", "Award Status", "Qualification Status", "Parent Email"],
            ["Nguyen Hoang Bach", "Ha Long Primary School", "3", "ICO", "Bronze", "Yes", "bach@example.com"],
            ["Tran Minh Khoa", "Le Quy Don", "4", "ICO", "Silver", "Yes", "khoa@example.com"],
        ]
    else:
        rows = [
            ["Nguyen Hoang Bach", "Ha Long Primary School", "3", "ICO", "Bronze", "Yes", "bach@example.com"],
            ["Tran Minh Khoa", "Le Quy Don", "4", "ICO", "Silver", "Yes", "khoa@example.com"],
        ]

    for row in rows:
        sheet_one.append(row)
        sheet_two.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_alias_merge_workbook_bytes() -> bytes:
    workbook = Workbook()
    primary = workbook.active
    primary.title = "IAIO_OUTPUT"
    copy_sheet = workbook.create_sheet("Bản sao IAIO Results")

    primary.append(
        [
            "Candidate's Name",
            "School/ Institution Name",
            "Grade",
            "Competition",
            "Award Status",
            "Qualification Status",
            "Parent Email",
        ]
    )
    primary.append(
        [
            "Le Ngoc Linh Dan",
            "Truong Tieu hoc Gia Thuong",
            "1",
            "IAIO",
            "Bronze",
            "Yes",
            "linhdan@example.com",
        ]
    )

    copy_sheet.append(["DANH SACH THI SINH", "", "", "", ""])
    copy_sheet.append(["STT", "Họ và tên thí sinh", "Trường/Đơn vị", "Khối lớp", "Tỉnh/Thành phố"])
    copy_sheet.append([1, "Le Ngoc Linh Dan", "Truong Tieu hoc Gia Thuong", "1", "TP. Ha Noi"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_dual_subject_workbook_bytes() -> bytes:
    workbook = Workbook()
    iaio_sheet = workbook.active
    iaio_sheet.title = "IAIO_OUTPUT"
    ico_sheet = workbook.create_sheet("ICO_OUTPUT")

    headers = [
        "Candidate's Name",
        "School/ Institution Name",
        "Grade",
        "Competition",
        "Award Status",
        "Qualification Status",
        "Parent Email",
    ]
    iaio_sheet.append(headers)
    ico_sheet.append(headers)
    iaio_sheet.append(["Nguyen Ky Anh", "TH Nam Trung Yen", "2", "IAIO", "Silver", "Yes", "iaio@example.com"])
    ico_sheet.append(["Nguyen Ky Anh", "TH Nam Trung Yen", "2", "ICO", "Gold", "Yes", "ico@example.com"])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TabularImportTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(username="tester", password="testpass123")
        self.client.force_authenticate(self.user)

    def test_import_rows_persists_participants_enrollments_and_results(self):
        competition = Competition.objects.create(name="Math Olympiad", academic_year="2025-2026")
        job = DataImportJob.objects.create(
            competition=competition,
            source_type=DataImportJob.SourceType.CSV,
            source_filename="students.csv",
        )

        summary = import_rows(
            competition=competition,
            rows=[
                {
                    "Student ID": "ST001",
                    "Student Name": "Nguyen Van A",
                    "Email": "a@example.com",
                    "School": "Le Hong Phong",
                    "Grade": "12",
                    "Subject": "Mathematics",
                    "Award": "Gold Medal",
                    "Certificate Code": "MATH-G-001",
                }
            ],
            import_job=job,
            imported_source="csv",
        )

        self.assertEqual(summary.row_count, 1)
        self.assertEqual(Participant.objects.count(), 1)
        self.assertEqual(CompetitionEnrollment.objects.count(), 1)
        self.assertEqual(CompetitionResult.objects.count(), 1)

        participant = Participant.objects.get()
        result = CompetitionResult.objects.get()

        self.assertEqual(participant.external_student_id, "ST001")
        self.assertEqual(result.certificate_code, "MATH-G-001")
        self.assertEqual(job.status, DataImportJob.Status.COMPLETED)

    def test_import_rows_supports_output_sheet_aliases_and_source_snapshot(self):
        competition = Competition.objects.create(name="SCO International Coding Olympiad", academic_year="2026-2027")
        job = DataImportJob.objects.create(
            competition=competition,
            source_type=DataImportJob.SourceType.XLSX,
            source_filename="ICO_OUTPUT.xlsx",
        )

        summary = import_rows(
            competition=competition,
            rows=[
                {
                    "Candidate's Name": "Nguyen Hoang Bach",
                    "School/ Institution Name": "Ha Long Primary School",
                    "Grade": "3",
                    "Competition": "ICO",
                    "Parent Email": "bach@example.com",
                    "Award Status": "Bronze",
                    "Qualification Status": "Yes",
                    "Registration Status": "Confirmed",
                }
            ],
            import_job=job,
            imported_source="xlsx",
        )

        self.assertEqual(summary.row_count, 1)
        enrollment = CompetitionEnrollment.objects.select_related("participant").get()
        result = CompetitionResult.objects.get()

        self.assertEqual(enrollment.participant.full_name, "Nguyen Hoang Bach")
        self.assertEqual(enrollment.participant.grade, "3")
        self.assertEqual(enrollment.subject, "ICO")
        self.assertEqual(enrollment.notes, "Confirmed")
        self.assertEqual(enrollment.source_data_json["Candidate's Name"], "Nguyen Hoang Bach")
        self.assertEqual(result.award, "Bronze")
        self.assertEqual(result.qualified_round, "Yes")

    def test_import_rows_rematches_existing_certificate_pages(self):
        competition = Competition.objects.create(name="Coding Olympiad", academic_year="2025-2026")
        batch = SourcePdfBatch.objects.create(
            competition=competition,
            original_filename="certificates.pdf",
            confirmed_competition_name=competition.name,
            competition_confirmation_status=SourcePdfBatch.ConfirmationStatus.CONFIRMED,
            status=SourcePdfBatch.Status.COMPLETED,
        )
        batch.uploaded_file.save("certificates.pdf", ContentFile(b"%PDF-1.4"), save=True)

        page = CertificatePage.objects.create(
            source_batch=batch,
            page_number=1,
            processing_status=CertificatePage.ProcessingStatus.EXTRACTED,
        )
        CertificateExtraction.objects.create(
            certificate_page=page,
            student_name="Nguyen Van A",
            normalized_student_name="nguyen van a",
            award="Gold Medal",
            certificate_code="MATH-G-001",
        )

        job = DataImportJob.objects.create(
            competition=competition,
            source_type=DataImportJob.SourceType.CSV,
            source_filename="students.csv",
        )

        import_rows(
            competition=competition,
            rows=[
                {
                    "Student ID": "ST001",
                    "Student Name": "Nguyen Van A",
                    "Email": "a@example.com",
                    "School": "Le Hong Phong",
                    "Grade": "12",
                    "Subject": "Mathematics",
                    "Award": "Gold Medal",
                    "Certificate Code": "MATH-G-001",
                }
            ],
            import_job=job,
            imported_source="csv",
        )

        page.refresh_from_db()
        match = page.match

        self.assertEqual(match.competition_enrollment.participant.full_name, "Nguyen Van A")
        self.assertEqual(match.confidence_label, "high")
        self.assertEqual(page.processing_status, CertificatePage.ProcessingStatus.MATCHED)
        self.assertEqual(job.details_json["rematched_pages"], 1)

    def test_import_file_inspect_returns_sheet_metadata(self):
        competition = Competition.objects.create(name="Inspect Workbook", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students.xlsx",
            build_workbook_bytes(include_header=True),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["import_job"]["status"], DataImportJob.Status.INSPECTED)
        self.assertEqual(payload["inspection"]["sheet_count"], 2)
        self.assertEqual(payload["inspection"]["sheets"][0]["name"], "ICO_OUTPUT")
        self.assertTrue(payload["inspection"]["sheets"][0]["detected_has_header"])

    def test_import_file_inspect_uses_column_letters_without_headers(self):
        competition = Competition.objects.create(name="Inspect Headerless Workbook", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students_no_header.xlsx",
            build_workbook_bytes(include_header=False),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )

        self.assertEqual(response.status_code, 201)
        sheet = response.json()["inspection"]["sheets"][0]
        self.assertFalse(sheet["detected_has_header"])
        self.assertEqual(sheet["columns"][:4], ["A", "B", "C", "D"])

    def test_import_file_execute_uses_shared_mapping_across_selected_sheets(self):
        competition = Competition.objects.create(name="Execute Workbook", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students.xlsx",
            build_workbook_bytes(include_header=True),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        inspect_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )
        import_job_id = inspect_response.json()["import_job"]["id"]

        execute_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/execute/",
            {
                "import_job_id": import_job_id,
                "selected_sheets": ["ICO_OUTPUT", "IAIO_OUTPUT"],
                "has_header": True,
                "mapping": {
                    "student_name": "Candidate's Name",
                    "school_name": "School/ Institution Name",
                    "grade": "Grade",
                    "competition": "Competition",
                    "award": "Award Status",
                    "qualified_round": "Qualification Status",
                    "email": "Parent Email",
                },
            },
            format="json",
        )

        self.assertEqual(execute_response.status_code, 200)
        self.assertEqual(Participant.objects.count(), 2)
        self.assertEqual(CompetitionEnrollment.objects.count(), 2)
        job = DataImportJob.objects.get(pk=import_job_id)
        self.assertEqual(job.status, DataImportJob.Status.COMPLETED)
        self.assertEqual(job.details_json["selected_sheets"], ["IAIO_OUTPUT", "ICO_OUTPUT"])

    def test_import_file_execute_accepts_column_letters_for_headerless_workbook(self):
        competition = Competition.objects.create(name="Execute Headerless Workbook", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students_no_header.xlsx",
            build_workbook_bytes(include_header=False),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        inspect_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )
        import_job_id = inspect_response.json()["import_job"]["id"]

        execute_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/execute/",
            {
                "import_job_id": import_job_id,
                "selected_sheets": ["ICO_OUTPUT"],
                "has_header": False,
                "mapping": {
                    "student_name": "A",
                    "school_name": "B",
                    "grade": "C",
                    "competition": "D",
                    "award": "E",
                    "qualified_round": "F",
                    "email": "G",
                },
            },
            format="json",
        )

        self.assertEqual(execute_response.status_code, 200)
        participant = Participant.objects.get(full_name="Nguyen Hoang Bach")
        enrollment = CompetitionEnrollment.objects.get(participant=participant)
        self.assertEqual(enrollment.source_sheet_name, "ICO_OUTPUT")
        self.assertEqual(enrollment.source_data_json["A"], "Nguyen Hoang Bach")
        self.assertEqual(enrollment.subject, "ICO")

    def test_import_file_execute_rejects_replaying_a_completed_job(self):
        competition = Competition.objects.create(name="Replay Guard Workbook", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students.xlsx",
            build_workbook_bytes(include_header=True),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        inspect_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )
        import_job_id = inspect_response.json()["import_job"]["id"]
        payload = {
            "import_job_id": import_job_id,
            "selected_sheets": ["ICO_OUTPUT"],
            "has_header": True,
            "mapping": {
                "student_name": "Candidate's Name",
                "school_name": "School/ Institution Name",
                "grade": "Grade",
                "competition": "Competition",
                "award": "Award Status",
                "qualified_round": "Qualification Status",
                "email": "Parent Email",
            },
        }

        first_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/execute/",
            payload,
            format="json",
        )
        replay_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/execute/",
            payload,
            format="json",
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(replay_response.status_code, 400)
        self.assertIn("no longer executable", replay_response.json()["detail"])

    def test_import_file_execute_merges_rows_across_selected_sheets_using_alias_headers(self):
        competition = Competition.objects.create(name="Workbook Merge", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students_merge.xlsx",
            build_alias_merge_workbook_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        inspect_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )
        import_job_id = inspect_response.json()["import_job"]["id"]

        execute_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/execute/",
            {
                "import_job_id": import_job_id,
                "selected_sheets": ["IAIO_OUTPUT", "Bản sao IAIO Results"],
                "has_header": True,
                "mapping": {
                    "student_name": "Candidate's Name",
                    "school_name": "School/ Institution Name",
                    "grade": "Grade",
                    "competition": "Competition",
                    "award": "Award Status",
                    "qualified_round": "Qualification Status",
                    "email": "Parent Email",
                },
            },
            format="json",
        )

        self.assertEqual(execute_response.status_code, 200)
        self.assertEqual(Participant.objects.count(), 1)
        self.assertEqual(CompetitionEnrollment.objects.count(), 1)
        enrollment = CompetitionEnrollment.objects.select_related("participant").get()
        self.assertEqual(enrollment.subject, "IAIO")
        self.assertEqual(enrollment.participant.grade, "1")
        self.assertEqual(enrollment.source_data_json["Tỉnh/Thành phố"], "TP. Ha Noi")

    def test_import_file_execute_keeps_same_student_in_multiple_subjects(self):
        competition = Competition.objects.create(name="Dual Subject Workbook", academic_year="2026-2027")
        upload = SimpleUploadedFile(
            "students_dual.xlsx",
            build_dual_subject_workbook_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        inspect_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/inspect/",
            {"file": upload},
        )
        import_job_id = inspect_response.json()["import_job"]["id"]

        execute_response = self.client.post(
            f"/api/competitions/{competition.id}/import-file/execute/",
            {
                "import_job_id": import_job_id,
                "selected_sheets": ["IAIO_OUTPUT", "ICO_OUTPUT"],
                "has_header": True,
                "mapping": {
                    "student_name": "Candidate's Name",
                    "school_name": "School/ Institution Name",
                    "grade": "Grade",
                    "competition": "Competition",
                    "award": "Award Status",
                    "qualified_round": "Qualification Status",
                    "email": "Parent Email",
                },
            },
            format="json",
        )

        self.assertEqual(execute_response.status_code, 200)
        self.assertEqual(Participant.objects.count(), 1)
        self.assertEqual(CompetitionEnrollment.objects.count(), 2)
        self.assertEqual(CompetitionResult.objects.count(), 2)
        self.assertEqual(
            sorted(CompetitionEnrollment.objects.values_list("subject", flat=True)),
            ["IAIO", "ICO"],
        )

    def test_delete_import_job_removes_imported_rows_and_resets_related_matches(self):
        competition = Competition.objects.create(name="Delete Import Job", academic_year="2026-2027", subject="ICO")
        job = DataImportJob.objects.create(
            competition=competition,
            source_type=DataImportJob.SourceType.XLSX,
            source_filename="delete_me.xlsx",
            status=DataImportJob.Status.COMPLETED,
        )

        import_rows(
            competition=competition,
            rows=[
                {
                    "Candidate's Name": "Nguyen Hoang Bach",
                    "School/ Institution Name": "Ha Long Primary School",
                    "Grade": "3",
                    "Competition": "ICO",
                    "Parent Email": "bach@example.com",
                    "Award Status": "Bronze",
                    "Qualification Status": "Yes",
                }
            ],
            import_job=job,
            imported_source="xlsx",
        )

        enrollment = CompetitionEnrollment.objects.select_related("participant").get()
        result = CompetitionResult.objects.get()
        batch = SourcePdfBatch.objects.create(
            competition=competition,
            original_filename="certificates.pdf",
            confirmed_competition_name=competition.name,
            competition_confirmation_status=SourcePdfBatch.ConfirmationStatus.CONFIRMED,
            status=SourcePdfBatch.Status.COMPLETED,
        )
        batch.uploaded_file.save("certificates.pdf", ContentFile(b"%PDF-1.4"), save=True)
        page = CertificatePage.objects.create(
            source_batch=batch,
            page_number=1,
            processing_status=CertificatePage.ProcessingStatus.MATCHED,
        )
        CertificateExtraction.objects.create(
            certificate_page=page,
            student_name="Nguyen Hoang Bach",
            normalized_student_name="nguyen hoang bach",
            school_name="Ha Long Primary School",
            normalized_school_name="ha long primary school",
            grade="3",
            award="Bronze",
            subject="ICO",
        )
        CertificateMatch.objects.create(
            certificate_page=page,
            competition_enrollment=enrollment,
            competition_result=result,
            confidence_score=100,
            confidence_label=CertificateMatch.ConfidenceLabel.HIGH,
            matched_by=CertificateMatch.MatchedBy.CERTIFICATE_CODE,
            requires_review=False,
            is_approved=True,
            rationale="Exact code",
        )

        response = self.client.delete(f"/api/import-jobs/{job.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(DataImportJob.objects.filter(pk=job.id).exists())
        self.assertEqual(CompetitionResult.objects.count(), 0)
        self.assertEqual(CompetitionEnrollment.objects.count(), 0)
        self.assertEqual(Participant.objects.count(), 0)
        page.refresh_from_db()
        page.match.refresh_from_db()
        self.assertFalse(page.match.is_approved)
        self.assertIsNone(page.match.competition_enrollment)
        self.assertEqual(page.processing_status, CertificatePage.ProcessingStatus.REVIEW_REQUIRED)

    def test_delete_import_job_preserves_rows_from_other_jobs(self):
        competition = Competition.objects.create(name="Delete One Import Job", academic_year="2026-2027")
        first_job = DataImportJob.objects.create(
            competition=competition,
            source_type=DataImportJob.SourceType.XLSX,
            source_filename="first.xlsx",
            status=DataImportJob.Status.COMPLETED,
        )
        second_job = DataImportJob.objects.create(
            competition=competition,
            source_type=DataImportJob.SourceType.XLSX,
            source_filename="second.xlsx",
            status=DataImportJob.Status.COMPLETED,
        )

        import_rows(
            competition=competition,
            rows=[
                {"Student Name": "First Student", "Email": "first@example.com", "Award": "Bronze"},
            ],
            import_job=first_job,
            imported_source="xlsx",
        )
        import_rows(
            competition=competition,
            rows=[
                {"Student Name": "Second Student", "Email": "second@example.com", "Award": "Silver"},
            ],
            import_job=second_job,
            imported_source="xlsx",
        )

        response = self.client.delete(f"/api/import-jobs/{first_job.id}/")

        self.assertEqual(response.status_code, 204)
        self.assertFalse(DataImportJob.objects.filter(pk=first_job.id).exists())
        self.assertTrue(DataImportJob.objects.filter(pk=second_job.id).exists())
        self.assertEqual(Participant.objects.count(), 1)
        self.assertEqual(CompetitionEnrollment.objects.count(), 1)
        self.assertEqual(CompetitionResult.objects.count(), 1)
        self.assertEqual(Participant.objects.get().full_name, "Second Student")
