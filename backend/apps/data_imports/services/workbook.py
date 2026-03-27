from __future__ import annotations

import importlib.util
from pathlib import Path

import pandas as pd
from django.core.exceptions import ValidationError
from openpyxl.utils import get_column_letter

from apps.common.text import infer_competition_code, normalize_grade_value, safe_get


PREVIEW_ROW_COUNT = 6
HEADER_HINTS = [
    "student id",
    "student name",
    "candidate name",
    "candidate's name",
    "school",
    "grade",
    "competition",
    "subject",
    "award",
    "certificate code",
    "qualification",
    "qualified round",
    "email",
    "notes",
    "registration status",
]

FIELD_ALIASES: dict[str, list[str]] = {
    "student_id": ["student id", "id", "student no", "candidate id", "source row"],
    "student_name": ["student name", "candidate's name", "candidate name", "full name", "name", "họ và tên thí sinh"],
    "email": ["email", "parent email", "mail", "teacher email"],
    "school_name": ["school", "school name", "school/ institution name", "organization", "institution", "trường/đơn vị"],
    "grade": ["grade", "class", "khối lớp"],
    "competition": ["competition", "competition code", "olympiad", "contest"],
    "award": ["award", "award status", "result", "prize"],
    "qualified_round": ["qualified round", "qualification status", "round"],
    "notes": ["notes", "registration status", "remarks", "comment"],
}


def _supports_xls() -> bool:
    return importlib.util.find_spec("xlrd") is not None


def _load_dataframe(file_path: Path, *, sheet_name: str | None = None) -> pd.DataFrame:
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(file_path, header=None, dtype=str, keep_default_na=False).fillna("")
    if suffix == ".xls" and not _supports_xls():
        raise ValidationError("Inspecting .xls files requires the xlrd package, which is not installed")
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            dtype=str,
            keep_default_na=False,
        ).fillna("")
    raise ValidationError(f"Unsupported import file: {file_path.name}")


def _load_workbook_frames(file_path: Path) -> list[tuple[str, pd.DataFrame]]:
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return [("Sheet1", _load_dataframe(file_path))]
    if suffix in {".xlsx", ".xls"}:
        if suffix == ".xls" and not _supports_xls():
            raise ValidationError("Inspecting .xls files requires the xlrd package, which is not installed")
        workbook = pd.ExcelFile(file_path)
        return [
            (sheet_name, _load_dataframe(file_path, sheet_name=sheet_name))
            for sheet_name in workbook.sheet_names
        ]
    raise ValidationError(f"Unsupported import file: {file_path.name}")


def _coerce_row(value_row: list) -> list[str]:
    return ["" if value is None else str(value).strip() for value in value_row]


def _frame_to_rows(frame: pd.DataFrame) -> list[list[str]]:
    if frame.empty:
        return []
    return [_coerce_row(row) for row in frame.values.tolist()]


def _alpha_ratio(values: list[str]) -> float:
    non_empty = [value for value in values if value]
    if not non_empty:
        return 0
    return sum(1 for value in non_empty if any(character.isalpha() for character in value)) / len(non_empty)


def detect_header_row_index(rows: list[list[str]]) -> int | None:
    if len(rows) < 2:
        return None

    max_scan = min(4, len(rows))
    for row_index in range(max_scan):
        current_row = [value.strip() for value in rows[row_index]]
        next_row = [value.strip() for value in rows[row_index + 1]] if row_index + 1 < len(rows) else []
        current_non_empty = [value for value in current_row if value]
        if len(current_non_empty) < 2:
            continue

        normalized_header_cells = [value.lower() for value in current_non_empty]
        header_hint_matches = sum(
            1
            for value in normalized_header_cells
            if any(hint in value for hint in HEADER_HINTS)
        )
        if header_hint_matches >= 2:
            return row_index

        if not next_row:
            continue

        unique_ratio = len({value.lower() for value in current_non_empty}) / len(current_non_empty)
        current_alpha_ratio = _alpha_ratio(current_non_empty)
        next_alpha_ratio = _alpha_ratio(next_row)

        if unique_ratio >= 0.8 and current_alpha_ratio >= 0.75 and current_alpha_ratio >= next_alpha_ratio + 0.25:
            return row_index

    return None


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    normalized_headers: list[str] = []
    for index, value in enumerate(headers, start=1):
        base = value.strip() or get_column_letter(index)
        counter = seen.get(base, 0) + 1
        seen[base] = counter
        normalized_headers.append(base if counter == 1 else f"{base} ({counter})")
    return normalized_headers


def build_positional_columns(column_count: int) -> list[str]:
    return [get_column_letter(index) for index in range(1, column_count + 1)]


def _resolved_value(
    source_row: dict[str, str],
    *,
    field_name: str,
    preferred_column: str,
    available_columns: list[str],
    sheet_name: str,
) -> str:
    if preferred_column and preferred_column in available_columns:
        value = source_row.get(preferred_column, "").strip()
    else:
        value = safe_get(source_row, *(FIELD_ALIASES.get(field_name, [])))

    if field_name == "competition" and not value:
        value = infer_competition_code(sheet_name)
    if field_name == "grade":
        value = normalize_grade_value(value)
    return value.strip()


def _sheet_inspection(sheet_name: str, rows: list[list[str]]) -> dict:
    column_count = max((len(row) for row in rows), default=0)
    positional_columns = build_positional_columns(column_count)
    header_row_index = detect_header_row_index(rows)
    detected_has_header = header_row_index is not None
    header_columns = _dedupe_headers(rows[header_row_index][:column_count]) if detected_has_header and rows else positional_columns
    data_start_index = (header_row_index + 1) if detected_has_header and header_row_index is not None else 0
    data_row_count = max(len(rows) - data_start_index, 0)
    preview_rows = rows[header_row_index:] if detected_has_header and header_row_index is not None else rows
    return {
        "name": sheet_name,
        "row_count": data_row_count,
        "column_count": column_count,
        "detected_has_header": detected_has_header,
        "columns": header_columns if detected_has_header else positional_columns,
        "header_columns": header_columns,
        "positional_columns": positional_columns,
        "preview_matrix": [row[:column_count] + [""] * max(column_count - len(row), 0) for row in preview_rows[:PREVIEW_ROW_COUNT]],
    }


def inspect_tabular_file(file_path: Path) -> dict:
    file_path = Path(file_path)
    sheets = [_sheet_inspection(sheet_name, _frame_to_rows(frame)) for sheet_name, frame in _load_workbook_frames(file_path)]
    return {
        "file_type": file_path.suffix.lower().lstrip("."),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def normalize_selected_sheets(
    file_path: Path,
    *,
    selected_sheets: list[str],
    has_header: bool,
    mapping: dict[str, str],
) -> list[dict]:
    file_path = Path(file_path)
    if not selected_sheets:
        raise ValidationError("Select at least one sheet to import")
    if not mapping.get("student_name"):
        raise ValidationError("Student Name mapping is required")

    workbook_frames = dict(_load_workbook_frames(file_path))
    prepared_rows: list[dict] = []

    for sheet_name in selected_sheets:
        if sheet_name not in workbook_frames:
            raise ValidationError(f"Sheet '{sheet_name}' is not available in the uploaded workbook")

        rows = _frame_to_rows(workbook_frames[sheet_name])
        column_count = max((len(row) for row in rows), default=0)
        positional_columns = build_positional_columns(column_count)
        header_row_index = detect_header_row_index(rows)
        effective_header_index = header_row_index if has_header and header_row_index is not None else 0
        header_columns = _dedupe_headers(rows[effective_header_index][:column_count]) if rows else positional_columns
        available_columns = header_columns if has_header else positional_columns

        data_start_index = (effective_header_index + 1) if has_header and rows else 0
        data_rows = rows[data_start_index:] if rows else []
        row_number_start = data_start_index + 1 if has_header else 1
        for offset, row in enumerate(data_rows):
            padded_row = row[:column_count] + [""] * max(column_count - len(row), 0)
            source_row = dict(zip(available_columns, padded_row))
            canonical_payload = {
                field_name: _resolved_value(
                    source_row,
                    field_name=field_name,
                    preferred_column=source_column,
                    available_columns=available_columns,
                    sheet_name=sheet_name,
                )
                for field_name, source_column in mapping.items()
                if source_column or field_name == "competition"
            }
            if not canonical_payload.get("student_name"):
                raise ValidationError(
                    f"Sheet '{sheet_name}' cannot resolve Student Name from the current mapping or known header aliases"
                )
            prepared_rows.append(
                {
                    "sheet_name": sheet_name,
                    "source_row_number": row_number_start + offset,
                    "sheet_row_key": str(row_number_start + offset),
                    "source_data_json": {
                        **source_row,
                        "_sheet_name": sheet_name,
                    },
                    "payload": canonical_payload,
                }
            )

    return prepared_rows
